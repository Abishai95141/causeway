"""
Document Extractor

Converts binary document formats (PDF, XLSX) to plain text.
Routes by content_type / file extension so downstream components
always receive text, never raw bytes.
"""

import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)


class DocumentExtractor:
    """
    Extracts text from binary document formats.

    Supported formats:
    - PDF  → pypdf (already in pyproject.toml)
    - XLSX → openpyxl (already in pyproject.toml)
    - MD   → pass-through (UTF-8 decode)
    - TXT  → pass-through (UTF-8 decode)
    """

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def extract(
        self,
        content: bytes,
        content_type: str,
        filename: str = "",
    ) -> str:
        """
        Extract text from raw bytes based on content type / extension.

        Args:
            content:      Raw file bytes.
            content_type: MIME type (e.g. "application/pdf").
            filename:     Original filename (used for extension fallback).

        Returns:
            Extracted plain-text string.
        """
        ext = self._extension(filename)

        if content_type == "application/pdf" or ext == "pdf":
            return self._extract_pdf(content)

        if (
            content_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or ext == "xlsx"
        ):
            return self._extract_xlsx(content)

        if content_type == "text/markdown" or ext == "md":
            return content.decode("utf-8", errors="replace")

        # text/plain and everything else — best-effort UTF-8
        return content.decode("utf-8", errors="replace")

    # ------------------------------------------------------------------ #
    # Private helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _extension(filename: str) -> str:
        """Return lowercased file extension without the dot."""
        if "." in filename:
            return filename.rsplit(".", 1)[-1].lower()
        return ""

    @staticmethod
    def _extract_pdf(content: bytes) -> str:
        """Extract text from a PDF using pypdf."""
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        pages: list[str] = []
        for idx, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {idx + 1}]\n{text}")
        if not pages:
            logger.warning("PDF extraction produced zero text — file may be scanned/image-only")
            return ""
        return "\n\n".join(pages)

    @staticmethod
    def _extract_xlsx(content: bytes) -> str:
        """Extract text from an Excel workbook using openpyxl."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets: list[str] = []
        for ws in wb.worksheets:
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                sheets.append(f"[Sheet: {ws.title}]\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets)
